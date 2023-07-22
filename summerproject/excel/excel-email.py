import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Open the Excel file
workbook = openpyxl.load_workbook('excel_email.xlsx')
worksheet = workbook['Sheet1']  # Replace 'Sheet1' with the name of your sheet

# Set up the SMTP server and login to your email account
smtp_server = 'smtp.gmail.com'  # Update with your SMTP server
smtp_port = 587  # Update with the appropriate port number
sender_email = 'kac2035@gmail.com'  # Update with your email address
sender_password = 'jgwqdoopraytdnzg'  # Update with your email password

# Iterate through each row in the Excel file
for row in worksheet.iter_rows(min_row=2, values_only=True):
    email = row[0]  # Column index of the email address in Excel (0-based)
    data = row[1]  # Column index of the data related to the email (0-based)

    # Create the email message
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = email
    message['Subject'] = 'Your Data'

    # Add the data to the email body
    message.attach(MIMEText(data, 'plain'))

    # Connect to the SMTP server and send the email
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(message)
        print(f"Email sent to {email}")

print("All emails sent successfully.")











/////////#

import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

# Open the Excel file
workbook = openpyxl.load_workbook('excel_email.xlsx')

# Get the email sheet
email_sheet = workbook['email']

# Set up the SMTP server and login to your Gmail account
smtp_server = 'smtp.gmail.com'
smtp_port = 587
sender_email = 'kac2035@gmail.com'  # Update with your email address
sender_password = 'jgwqdoopraytdnzg'  # Update with your email password

# Iterate through each row in the email sheet
for email_row in email_sheet.iter_rows(min_row=2, values_only=True):
    email = email_row[1]  # Column index of the email address in the email sheet (0-based)

    # Iterate through each sheet except the email sheet
    for sheet_name in workbook.sheetnames:
        if sheet_name == 'email':
            continue  # Skip the email sheet

        worksheet = workbook[sheet_name]

        # Create the email message
        message = MIMEMultipart()
        message['From'] = sender_email
        message['To'] = email
        message['Subject'] = f'Data Sheet - {sheet_name}'

        # Create a new workbook for the sheet and copy its content
        sheet_workbook = openpyxl.Workbook()
        sheet_copy = sheet_workbook.active
        sheet_copy.title = 'Data'
        for row in worksheet.iter_rows(values_only=True):
            sheet_copy.append(row)

        # Save the sheet workbook as 'data_sheet.xlsx'
        file_path = f'data_sheet_{sheet_name}.xlsx'
        sheet_workbook.save(file_path)

        # Attach the sheet to the email
        with open(file_path, 'rb') as file:
            attachment = MIMEApplication(file.read(), _subtype='xlsx')
            attachment.add_header('Content-Disposition', 'attachment', filename=f'data_sheet_{sheet_name}.xlsx')
            message.attach(attachment)

        # Connect to the SMTP server and send the email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(message)
            print(f"Email sent to {email} with the {sheet_name} sheet attached")

print("All emails sent successfully.")
