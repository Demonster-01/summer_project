import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

# Open the Excel file
workbook = openpyxl.load_workbook('excel_email.xlsx')

# Get the email sheet
email_sheet = workbook['email']

# Set up the SMTP server and login to your Gmail account
smtp_server = 'smtp.gmail.com'
smtp_port = 587
sender_email = 'kac2035@gmail.com'  # Update with your email address
sender_password = 'jgwqdoopraytdnzg'  # Update with your email password

# Create a dictionary to store email-sheet mappings
email_sheet_map = {}

# Iterate through each sheet except the email sheet
for sheet_name in workbook.sheetnames:
    if sheet_name == 'email':
        continue  # Skip the email sheet

    worksheet = workbook[sheet_name]

    # Add the email and sheet mapping to the dictionary
    for row in email_sheet.iter_rows(min_row=2, values_only=True):
        email = row[1]  # Column index of the email address in the email sheet (0-based)
        email_sheet_map[email] = sheet_name

# Iterate through each email in the email-sheet mapping
for email in email_sheet_map:
    sheet_name = email_sheet_map[email]
    worksheet = workbook[sheet_name]

    # Create the email message
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = email
    message['Subject'] = f'Data Sheet - {sheet_name}'

    # Create a new workbook for the sheet and copy its content
    sheet_workbook = openpyxl.Workbook()
    sheet_copy = sheet_workbook.active
    sheet_copy.title = 'Data'
    for row in worksheet.iter_rows(values_only=True):
        sheet_copy.append(row)

    # Save the sheet workbook as 'data_sheet.xlsx'
    file_path = f'{sheet_name}.xlsx'
    sheet_workbook.save(file_path)

    # Attach the sheet to the email
    with open(file_path, 'rb') as file:
        attachment = MIMEApplication(file.read(), _subtype='xlsx')
        attachment.add_header('Content-Disposition', 'attachment', filename=f'data_sheet_{sheet_name}.xlsx')
        message.attach(attachment)

    # Connect to the SMTP server and send the email
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(message)
        print(f"Email sent to {email} with the {sheet_name} sheet attached")

print("All emails sent successfully.")
